import networkx as nx
import matplotlib.pyplot as plt
from networkx.drawing.nx_agraph import graphviz_layout
from z3 import *
import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

def plot_dag_with_networkx(dag, pic_name='dag'):
    """
    Visualizes a Directed Acyclic Graph (DAG) and saves it as an image.

    Parameters:
    - dag: The DAG object to be plotted.
    - pic_name: The name of the output image file (default is 'dag').
    """
    G = nx.DiGraph()

    # Add nodes and edges
    for node_name in dag.nodes:
        node = dag.nodes[node_name]
        G.add_node(node_name, node_type=node.node_type)
        for child in node.children:
            G.add_edge(node_name, child.name)

    # Define node colors
    node_colors = {
        "INPUT": "#90EE90",  # Light green for input nodes
        "OUTPUT": "#FFB6C1",  # Light red for output nodes
    }

    # Assign colors to nodes
    node_color = [node_colors.get(G.nodes[node]["node_type"], "#87CEEB") for node in G.nodes]
    labels = {node: node for node in G.nodes}

    # Use Graphviz for layout
    pos = graphviz_layout(G, prog="dot", args="-Grankdir=LR -Granksep=20")

    # Adjust figure size based on the number of nodes
    num_nodes = len(G.nodes())
    fig_size = (num_nodes / 2, num_nodes / 3)

    plt.figure(figsize=fig_size)

    nx.draw(G, pos, with_labels=False, node_size=2000, node_color=node_color, edge_color='gray', arrows=True)

    nx.draw_networkx_labels(G, pos, labels, font_size=10, font_color='black')

    # Add node type labels
    for node, data in G.nodes(data=True):
        if data["node_type"] not in ["INPUT", "OUTPUT"]:
            x, y = pos[node]
            plt.text(
                x, y - 10,  # Offset by 10 units downward
                data["node_type"],
                fontsize=8,
                color='#d4342f',  # Red text
                horizontalalignment='center',
                verticalalignment='top'
            )

    plt.title(pic_name)
    plt.axis('off')
    plt.savefig(f'{pic_name}.png', dpi=300)
    plt.show()


def build_expression(output, inputs, adj):
    """
    Recursively constructs a Boolean logic expression for the given output node.

    Parameters:
    - output: The name of the output node.
    - inputs: A dictionary of input values.
    - adj: The adjacency list representing the DAG.

    Returns:
    - A Boolean expression representing the logic operation.
    """
    if output in inputs.keys():
        return inputs.get(output, None)

    expr_info = adj.get(output)
    if expr_info is None:
        return inputs.get(output, None)

    op, *args = expr_info
    if op is None or op.upper() == 'INPUT':
        return inputs.get(output, None)

    if op.upper() == 'OUTPUT':
        return build_expression(args[0], inputs, adj)

    subop = [build_expression(arg, inputs, adj) for arg in args]

    if op.upper() == 'AND':
        return And(*subop)
    elif op.upper() == 'OR':
        return Or(*subop)
    elif op.upper() == 'XOR':
        return Xor(*subop)
    elif op.upper() == 'NOT':
        if len(subop) != 1:
            raise ValueError(f"NOT operation must have exactly one argument, got {len(subop)}")
        return Not(subop[0])
    else:
        raise ValueError(f"Unknown operation {op} in node {output}")


def get_case_single(output_name, inputs, adj):
    """
    Generates all possible input cases that satisfy the output condition.

    Parameters:
    - output_name: The output node name.
    - inputs: A dictionary of input nodes.
    - adj: The adjacency list of the DAG.

    Returns:
    - A list of dictionaries containing input cases.
    """
    variance = {name: Bool(name) for name in inputs}
    true_expression = build_expression(output_name, variance, adj)
    false_expression = Not(true_expression)
    expressions = [(true_expression, output_name, 1), (false_expression, output_name, 0)]
    # print("Constructed logic expressions:", expressions)
    return get_solutions_z3(expressions, variance)


def get_solutions_z3(expressions, variance):
    """
    Uses Z3 solver to find all possible solutions for the given Boolean expressions.

    Parameters:
    - expressions: A list of Boolean expressions and expected outputs.
    - variance: A dictionary mapping variable names to Z3 Boolean variables.

    Returns:
    - A list of dictionaries representing all valid input cases.
    """

    solutions = []
    for expr, output_name, out in expressions:
        simplified_expr = simplify(expr)
        s = Solver()
        s.add(simplified_expr)
        sol = []
        while s.check() == sat:
            model = s.model()
            block = []
            model_set = [(v, is_true(model[v])) for v in model.decls()]
            for var in variance:
                va = variance[var]
                if model[va]:
                    block.append(va)
                else:
                    block.append(Not(va))
            s.add(Not(And(block)))
            if any(set(sol_set).issubset(model_set) for sol_set in sol):
                # s.add(Not(And([v if is_true(model[v]) else Not(v) for v in variance.values()])))
                continue
            sol.append(model_set)


        print(f'\nAll solutions that meet the condition that {output_name} is {out}: {expr}')

        for s in sol:
            formatted_sol = {v.name(): (1 if val else 0) if val is not None else '' for v, val in s}
            solutions.append(formatted_sol)
            print(formatted_sol)

    return solutions


def remove_included_dicts(dict_list):
    dict_list.sort(key=len)

    to_remove = set()
    for i, dict1 in enumerate(dict_list):
        for j in range(i + 1, len(dict_list)):
            dict2 = dict_list[j]
            if dict1.items() <= dict2.items():  # dict1 被 dict2 包含
                to_remove.add(i)
                break


    return [d for i, d in enumerate(dict_list) if i not in to_remove]


def get_case_in_multiple(dag):
    """
    Generates all possible input cases for multiple output nodes in the DAG.

    Parameters:
    - dag: The DAG object.

    Returns:
    - A list of unique input cases.
    """
    out_nodes =  [node.name for node in dag.get_outnode()]
    input_nodes = [dag.get_innode(out) for out in out_nodes]
    all_inputs = [node.name for node in dag.get_all_inputs()]
    cases = []

    for out_node, input_node in zip(out_nodes, input_nodes):
        case = get_case_single(out_node, input_node, dag.adj)
        cases += case
    #remove repeat case
    cases = remove_included_dicts(cases)
    unique_cases = set()
    for case in cases:
        missing_keys = [input for input in all_inputs if input not in case]

        for key in missing_keys:
            case[key] = random.choice([0, 1])  # Randomly assign missing inputs
        unique_cases.add(frozenset(case.items()))

    print(f'All unique cases: {len(unique_cases)}')
    in_cases = [dict(case) for case in unique_cases]

    return in_cases


def create_out(cases, dag):
    """
    Computes the output for each case based on the DAG.

    Parameters:
    - cases: A list of input cases.
    - dag: The DAG object.

    Returns:
    - A list of cases including both inputs and corresponding outputs.
    """
    new_cases = []
    for case in cases:
        out_nodes = dag.evaluate(case)
        merge_case = {**case, **out_nodes}
        new_cases.append(merge_case)

    print(f'All cases: {len(new_cases)}')
    for case in new_cases:
        print(dict(case))

    return new_cases


def save_cases(cases, dag_name='dag'):
    """
    Saves the generated test cases to a CSV file.

    Parameters:
    - cases: The list of test cases.
    - dag_name: The name of the DAG (used as the file name).
    """
    pd_cases = pd.DataFrame(cases)
    pd_cases.to_csv(f'{dag_name}.csv', index=False)


def compute_levels(dag):
    all_nodes = dag.topological_sort()
    levels = {node.name: 0 for node in all_nodes}
    for node in all_nodes:
        children = node.children
        for child in children:
            levels[child.name] =max(levels[child.name], levels[node.name]+1)
    return levels

def extract_path(levels,input_nodes):
    max_level = max(levels.values())
    columns = [f'Gate {i}' for i in range(max_level + 1)]
    paths = []
    for inp in input_nodes:
        path = [None] * (max_level + 1)
        current_node = inp
        path[levels[current_node.name]] = current_node.name  # 输入节点层级为0

        # 追踪路径直到输出
        while True:
            next_nodes = [node for node in current_node.children ]
            if not next_nodes:
                break
            next_node = next_nodes[0]  # 假设每个节点只有一个输出
            path[levels[next_node.name]] = next_node.name
            # if next_node.node_type=='OUTPUT':
            #     path[levels[next_node.name]] = next_node.name
            current_node = next_node

        paths.append(path)
    return paths, columns




def save_to_csv(paths, columns, gate_map):
    """
    Saves logic table data into an Excel file with formatted structure.

    :param paths: Data for the logic table (list of lists).
    :param columns: List of column names.
    :param gate_map: A dictionary mapping node names to logic gate types, e.g.,
                     {"G1": "AND", "G2": "OR", "G3": "NOT"}.
    """
    # Rename the first column to 'Input' and the last column to 'Output'
    columns[0] = 'Input'
    columns[-1] = 'Output'

    wb = Workbook()
    ws = wb.active

    # Ensure all column headers are strings
    columns = [str(col) for col in columns]

    # Identify the range of "Gate X" columns
    gate_indices = [i for i, col in enumerate(columns) if "Gate" in col]
    gate_start = gate_indices[0] if gate_indices else None
    gate_end = gate_indices[-1] if gate_indices else None

    # Write column headers in the second row
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=2, column=col_idx, value=col_name)

    # Merge cells in the first row for "Gate Type" header
    if gate_start is not None and gate_end is not None:
        ws.merge_cells(start_row=1, start_column=gate_start + 1,
                       end_row=1, end_column=gate_end + 1)
        ws.cell(row=1, column=gate_start + 1, value="Gate Type").alignment = Alignment(horizontal="center", vertical="center")

    # Write data (starting from the third row)
    for row_idx, path in enumerate(paths, start=3):
        for col_idx, value in enumerate(path, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Merge consecutive identical node names (Gate X) in each column
    for col_idx in range(1, len(columns) + 1):  # Iterate through all columns, including the last one
        current_value = None
        start_row = 3  # Data starts from row 3
        for row_idx in range(3, len(paths) + 3):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value == current_value:
                continue  # Skip if the value is the same as the previous row
            else:
                # Merge previous consecutive cells if they had the same value
                if current_value is not None and start_row < row_idx - 1:
                    ws.merge_cells(start_row=start_row, start_column=col_idx,
                                   end_row=row_idx - 1, end_column=col_idx)
                    for r in range(start_row, row_idx):
                        ws.cell(r, col_idx).alignment = Alignment(horizontal='center', vertical='center')
                current_value = cell_value
                start_row = row_idx

        # Merge the last block of identical values (including the Output column)
        if current_value is not None and start_row < len(paths) + 3:
            ws.merge_cells(start_row=start_row, start_column=col_idx,
                           end_row=len(paths) + 2, end_column=col_idx)
            for r in range(start_row, len(paths) + 3):
                ws.cell(r, col_idx).alignment = Alignment(horizontal='center', vertical='center')

    # Replace node names with their corresponding logic gate types
    for row_idx in range(3, len(paths) + 3):
        for col_idx in gate_indices:
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            if cell_value in gate_map:
                ws.cell(row=row_idx, column=col_idx + 1, value=gate_map[cell_value])

    # Adjust column widths for better readability
    for col_idx in range(1, ws.max_column + 1):
        max_length = max((len(str(ws.cell(row=row, column=col_idx).value)) if ws.cell(row=row, column=col_idx).value else 0)
                         for row in range(1, ws.max_row + 1))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = adjusted_width

    return wb

def save_dag_to_csv(dag,filename='dag_flow'):
    levels = compute_levels(dag)
    input_nodes = dag.get_all_inputs()
    paths, columns=extract_path(levels,input_nodes)
    logic_gates = {node.name: node.node_type for node in dag.nodes.values() if node.node_type!= 'OUTPUT' and node.node_type != 'INPUT'}

    wb = save_to_csv(paths,columns,logic_gates)
    wb.save(filename+'.xlsx')

    print(f'save dag to excel: {filename}')